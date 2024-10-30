import openpyxl

class ExcelReader:
    def __init__(self, file_path):
        self.workbook = openpyxl.load_workbook(file_path)

    def list_sheets(self):
        return self.workbook.sheetnames

    def get_data(self, sheet_name, row, field):
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Worksheet '{sheet_name}' does not exist in the workbook.")
        sheet = self.workbook[sheet_name]
        col = self.get_column_index(sheet, field)
        return sheet.cell(row=row, column=col).value

    def get_column_index(self, sheet, field):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == field:
                return col
        raise ValueError(f"Field '{field}' not found in the sheet.")

excel = ExcelReader('C://Users//gaurav//Desktop//Automation//CRM_Project//CRM_testdata.xlsx')  # Ensure the path is correct
print(excel.list_sheets())

